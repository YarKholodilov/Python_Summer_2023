import openpyxl
wb = openpyxl.load_workbook('hometask_10_1.xlsx')
ws = wb.active

dic = {}

for i in range(1, ws.max_row + 1):
    person = ws.cell(i, 1).value.title()
    production = int(ws.cell(i, 2).value)
    dic[person] = dic.get(person, 0) + production

wb.create_sheet('Total')
wb.active = wb['Total']

dic['ИТОГО:'] = dic.get('ИТОГО:', sum(dic.values()))
print(dic)
i = 1
for k, v in dic.items():
    wb['Total'].cell(i, 1).value = k
    wb['Total'].cell(i, 2).value = v
    i += 1

wb.save('hometask_10_1.xlsx')