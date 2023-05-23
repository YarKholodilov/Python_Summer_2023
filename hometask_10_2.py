import openpyxl
wb = openpyxl.load_workbook('hometask_10_2.xlsx')
wb.active = wb['april']
dic = {}

for i in range(1, wb['april'].max_row + 1):
    person = wb['april'].cell(i, 1).value.title()
    production = int(wb['april'].cell(i, 2).value)
    dic[person] = dic.get(person, 0) + production

wb.active = wb['may']
for i in range(1, wb['may'].max_row + 1):
    person = wb['may'].cell(i, 1).value.title()
    production = int(wb['may'].cell(i, 2).value)
    dic[person] = dic.get(person, 0) + production

wb.create_sheet('Total')
wb.active = wb['Total']

i = 1
for k, v in sorted(dic.items()):
    wb['Total'].cell(i, 1).value = k
    wb['Total'].cell(i, 2).value = v
    i += 1

wb['Total'].cell(len(dic)+1, 1).value = 'ИТОГО'
wb['Total'].cell(len(dic)+1, 2).value = sum(dic.values())

wb.save('hometask_10_2.xlsx')