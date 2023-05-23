import openpyxl
wb = openpyxl.load_workbook('hometask_10_3.xlsx')
wb.active = wb['production']

efficiency = list()

for i in range(1, wb['production'].max_row + 1):
    efficiency.append(wb['production'].cell(i, 2).value)

wb.create_sheet('analysis')
wb.active = wb['analysis']

wb['analysis'].cell(1, 1).value = 'Миниальное значение:'
wb['analysis'].cell(1, 2).value = min(efficiency)
wb['analysis'].cell(2, 1).value = 'Максимальное значение:'
wb['analysis'].cell(2, 2).value = max(efficiency)
wb['analysis'].cell(3, 1).value = 'Среднее арифметическое:'
wb['analysis'].cell(3, 2).value = sum(efficiency)/len(efficiency)
wb['analysis'].cell(4, 1).value = 'Медиана:'

if len(efficiency) % 2 != 0:
    x = (len(efficiency) // 2)
    mediana = efficiency[x]
else:
    x = (len(efficiency) // 2)
    y = (len(efficiency) // 2) - 1
    mediana = (efficiency[x] + efficiency[y]) / 2

wb['analysis'].cell(4, 2).value = mediana

wb.save('hometask_10_3.xlsx')